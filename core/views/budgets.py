"""
Quadro Orçamentário — Finanças 360
====================================
Orçamento por mês: cada conta tem uma meta mensal (amount) por mês do ano.
A view budget_dashboard exibe os 12 meses lado a lado com Realizado vs Meta.
"""
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
import calendar
import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from django.db.models.functions import Coalesce
from django.db.models import Value, DecimalField
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from ..models import Company, ChartOfAccounts, Transaction, Budget
from ..permissions import (
    role_required, has_role,
    ADMIN, GESTOR, MANAGERS, NON_PLANTOES,
)


MONTH_NAMES_SHORT = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun',
                     'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
MONTH_NAMES_FULL  = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                     'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
MONTHS = list(range(1, 13))


def _parse_year(value, default=None):
    if default is None:
        default = date.today().year
    try:
        digits = re.sub(r'[\.,\s]', '', str(value))
        year = int(digits)
        if 1900 <= year <= 2200:
            return year
    except (ValueError, TypeError):
        pass
    return default


def _build_budget_data(company, year):
    """
    Constrói a estrutura completa do quadro orçamentário.
    Retorna report_data (lista de dicts com hierarquia) + summary_data.

    Queries:
      - 1x ChartOfAccounts
      - 1x Budget (todos os meses do ano)
      - 1x Transaction (todos os lançamentos do ano)
    """
    # ── 1. Carrega contas ────────────────────────────────────────────
    accounts = list(ChartOfAccounts.objects.filter(company=company)
                    .select_related('parent_account').order_by('code'))
    account_map = {a.id: a for a in accounts}

    # ── 2. Carrega metas mensais: {account_id: {month: amount}} ─────
    budgets_qs = Budget.objects.filter(account__company=company, year=year)
    budget_raw = defaultdict(lambda: defaultdict(Decimal))
    for b in budgets_qs:
        budget_raw[b.account_id][b.month] = b.amount

    # ── 3. Carrega realizados: {account_id: {month: total}} ─────────
    actuals_qs = (
        Transaction.objects
        .filter(company=company, date__year=year)
        .values('account_id', 'date__month')
        .annotate(total=Coalesce(Sum('amount'), Value(Decimal('0.00'))))
    )
    actual_raw = defaultdict(lambda: defaultdict(Decimal))
    for row in actuals_qs:
        actual_raw[row['account_id']][row['date__month']] = row['total']

    # ── 4. Monta árvore e acumula totais ─────────────────────────────
    tree = {}
    for acc in accounts:
        tree[acc.id] = {
            'account': acc,
            'budget': [budget_raw[acc.id].get(m, Decimal('0')) for m in MONTHS],
            'actual': [actual_raw[acc.id].get(m, Decimal('0')) for m in MONTHS],
            'children': [],
            'level': 0,
        }

    root_ids = []
    for acc in accounts:
        if acc.parent_account_id and acc.parent_account_id in tree:
            tree[acc.parent_account_id]['children'].append(acc.id)
        else:
            root_ids.append(acc.id)

    def _aggregate(aid):
        node = tree[aid]
        for child_id in node['children']:
            _aggregate(child_id)
            child = tree[child_id]
            node['budget'] = [b + cb for b, cb in zip(node['budget'], child['budget'])]
            node['actual'] = [a + ca for a, ca in zip(node['actual'], child['actual'])]

    for rid in root_ids:
        _aggregate(rid)

    # ── 5. Calcula variações e status ─────────────────────────────────
    def _status(acc_type, budget_m, actual_m):
        """Retorna classe CSS de status para a célula do realizado."""
        if budget_m == 0:
            return 'neutral'
        if acc_type in ('D', 'E'):
            if actual_m > budget_m:
                return 'over'     # acima da meta → ruim para despesa
            elif actual_m > 0:
                return 'ok'
        else:  # Receita
            if actual_m >= budget_m:
                return 'ok'       # atingiu ou superou → bom
            elif actual_m > 0:
                return 'under'    # abaixo da meta → ruim para receita
        return 'neutral'

    def _build_list(aid, level=0):
        rows = []
        node = tree[aid]
        node['level'] = level
        monthly = []
        for i, m in enumerate(MONTHS):
            b = node['budget'][i]
            a = node['actual'][i]
            prev_a = node['actual'][i - 1] if i > 0 else Decimal('0')
            var_pct = ((a - prev_a) / prev_a * 100) if prev_a != 0 else None
            monthly.append({
                'budget': b,
                'actual': a,
                'var_pct': var_pct,
                'status': _status(node['account'].account_type, b, a),
            })
        node['monthly'] = monthly
        node['total_actual'] = sum(node['actual'])
        node['avg_budget'] = sum(node['budget']) / 12 if any(node['budget']) else Decimal('0')
        rows.append(node)
        for child_id in sorted(node['children'],
                               key=lambda x: tree[x]['account'].code):
            rows.extend(_build_list(child_id, level + 1))
        return rows

    report_data = []
    for rid in sorted(root_ids, key=lambda x: tree[x]['account'].code):
        report_data.extend(_build_list(rid))

    # ── 6. Summary (receitas / despesas / resultado) ────────────────
    def _find_root_by_code(code):
        for rid in root_ids:
            if tree[rid]['account'].code == code:
                return tree[rid]
        return None

    rev_node = _find_root_by_code('1')
    exp_node = _find_root_by_code('2')

    rev_monthly  = rev_node['actual'] if rev_node else [Decimal('0')] * 12
    exp_monthly  = exp_node['actual'] if exp_node else [Decimal('0')] * 12
    res_monthly  = [r - e for r, e in zip(rev_monthly, exp_monthly)]

    summary = {
        'revenue':       rev_monthly,
        'expense':       exp_monthly,
        'result':        res_monthly,
        'total_revenue': sum(rev_monthly),
        'total_expense': sum(exp_monthly),
        'total_result':  sum(res_monthly),
    }

    return report_data, summary


# ══════════════════════════════════════════════════════════════════
#  QUADRO ORÇAMENTÁRIO (dashboard)
# ══════════════════════════════════════════════════════════════════

@login_required
def budget_dashboard(request):
    # Analista de Plantões não acessa o quadro orçamentário
    if not has_role(request.user, *NON_PLANTOES):
        messages.error(request, "Você não tem permissão para acessar o quadro orçamentário.")
        return redirect('core:company_list')

    active_company_id = request.session.get('active_company_id')
    if not active_company_id:
        messages.error(request, "Selecione uma empresa primeiro.")
        return redirect('core:company_list')
    active_company = get_object_or_404(Company, pk=active_company_id, users=request.user)

    now_year = date.today().year
    year = _parse_year(request.GET.get('year', ''), default=now_year)

    # Anos disponíveis para o seletor
    tx_years  = set(Transaction.objects.filter(company=active_company)
                    .dates('date', 'year').values_list('date__year', flat=True))
    bud_years = set(Budget.objects.filter(account__company=active_company)
                    .values_list('year', flat=True).distinct())
    available_years = sorted((tx_years | bud_years | {now_year}), reverse=True)

    report_data, summary = _build_budget_data(active_company, year)

    context = {
        'company': active_company,
        'year': year,
        'available_years': available_years,
        'report_data': report_data,
        'summary': summary,
        'month_names': MONTH_NAMES_SHORT,
        'months': MONTHS,
        'can_manage': has_role(request.user, ADMIN, GESTOR),
    }
    return render(request, 'core/budget_dashboard.html', context)


# ══════════════════════════════════════════════════════════════════
#  LANÇAMENTO DE ORÇAMENTO (admin only)
# ══════════════════════════════════════════════════════════════════

@login_required
def budget_entry(request):
    """
    Página de lançamento de metas mensais por conta.
    Restrita a Admin e Gestor.
    """
    if not has_role(request.user, ADMIN, GESTOR):
        messages.error(request, "Você não tem permissão para lançar orçamentos.")
        return redirect('core:budget_dashboard')

    active_company_id = request.session.get('active_company_id')
    if not active_company_id:
        messages.error(request, "Selecione uma empresa primeiro.")
        return redirect('core:company_list')
    active_company = get_object_or_404(Company, pk=active_company_id, users=request.user)

    now_year = date.today().year
    year = _parse_year(request.GET.get('year', ''), default=now_year)

    # Contas folha (sem filhos) — únicas editáveis
    all_accounts = list(ChartOfAccounts.objects.filter(company=active_company)
                        .select_related('parent_account').order_by('code'))
    account_map = {a.id: a for a in all_accounts}

    leaf_ids = set(a.id for a in all_accounts
                   if not any(x.parent_account_id == a.id for x in all_accounts))

    # Carrega metas existentes para o ano
    existing = {
        (b.account_id, b.month): b
        for b in Budget.objects.filter(account__company=active_company, year=year)
    }

    if request.method == 'POST':
        to_create, to_update = [], []
        for acc in all_accounts:
            if acc.id not in leaf_ids:
                continue
            for m in MONTHS:
                raw = request.POST.get(f'budget_{acc.id}_{m}', '').strip()
                try:
                    value = Decimal(raw.replace(',', '.')) if raw else Decimal('0')
                except Exception:
                    value = Decimal('0')

                key = (acc.id, m)
                if key in existing:
                    obj = existing[key]
                    if obj.amount != value:
                        obj.amount = value
                        to_update.append(obj)
                else:
                    to_create.append(Budget(
                        company=active_company, account=acc,
                        year=year, month=m, amount=value,
                    ))

        if to_create:
            Budget.objects.bulk_create(to_create, batch_size=500, ignore_conflicts=True)
        if to_update:
            Budget.objects.bulk_update(to_update, ['amount'], batch_size=500)

        messages.success(request, f"Orçamento {year} salvo com sucesso!")
        return redirect(f"{request.path}?year={year}")

    # Monta tabela para exibição
    # Para contas pai: mostra somatório dos filhos em cada mês (read-only)
    def _get_children_ids(aid):
        direct = [a.id for a in all_accounts if a.parent_account_id == aid]
        result = list(direct)
        for child_id in direct:
            result.extend(_get_children_ids(child_id))
        return result

    def _sum_children(aid, m):
        children = _get_children_ids(aid)
        return sum(
            (existing.get((cid, m), None) or Budget(amount=Decimal('0'))).amount
            for cid in children if cid in leaf_ids
        )

    def _level(acc):
        lvl, p = 0, acc.parent_account
        while p:
            lvl += 1
            p = p.parent_account
        return lvl

    table_rows = []
    for acc in all_accounts:
        is_leaf = acc.id in leaf_ids
        level   = _level(acc)
        row = {
            'account': acc,
            'level': level,
            'is_leaf': is_leaf,
            'months': [],
        }
        for m in MONTHS:
            if is_leaf:
                val = existing.get((acc.id, m), None)
                row['months'].append(val.amount if val else Decimal('0'))
            else:
                row['months'].append(_sum_children(acc.id, m))
        row['annual_total'] = sum(row['months'])
        table_rows.append(row)

    # Anos disponíveis
    bud_years = set(Budget.objects.filter(account__company=active_company)
                    .values_list('year', flat=True).distinct())
    available_years = sorted((bud_years | {now_year, now_year + 1}), reverse=True)

    context = {
        'company': active_company,
        'year': year,
        'prev_year': year - 1,
        'next_year': year + 1,
        'available_years': available_years,
        'table_rows': table_rows,
        'month_names': MONTH_NAMES_SHORT,
        'months': MONTHS,
    }
    return render(request, 'core/budget_entry.html', context)


# ══════════════════════════════════════════════════════════════════
#  EXPORTAR EXCEL
# ══════════════════════════════════════════════════════════════════

@login_required
@role_required(*NON_PLANTOES)
def export_budget_xls(request):
    active_company_id = request.session.get('active_company_id')
    if not active_company_id:
        return redirect('core:company_list')
    active_company = get_object_or_404(Company, pk=active_company_id, users=request.user)

    year = _parse_year(request.GET.get('year', ''))
    report_data, _ = _build_budget_data(active_company, year)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Orçamento {year}"

    # Estilos
    hdr_font   = Font(bold=True, color='FFFFFF')
    hdr_fill   = PatternFill(start_color='1A56DB', end_color='1A56DB', fill_type='solid')
    ok_fill    = PatternFill(start_color='198754', end_color='198754', fill_type='solid')
    over_fill  = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    under_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
    wht_font   = Font(color='FFFFFF', bold=True)
    blk_font   = Font(color='000000', bold=True)

    # Cabeçalho
    headers = ['Conta', 'Tipo']
    for mn in MONTH_NAMES_SHORT:
        headers += [f'{mn} Meta', f'{mn} Real']
    headers += ['Total Realizado', 'Média Meta']
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')

    for row in report_data:
        acc = row['account']
        indent = '    ' * row['level']
        cells = [
            f"{indent}{acc.code} — {acc.name}",
            'Receita' if acc.account_type == 'R' else 'Despesa',
        ]
        for md in row['monthly']:
            cells += [float(md['budget']), float(md['actual'])]
        cells += [float(row['total_actual']), float(row['avg_budget'])]
        ws.append(cells)

        cur_row = ws.max_row
        for i, md in enumerate(row['monthly']):
            real_col = 3 + i * 2 + 1  # coluna do Realizado
            cell = ws.cell(row=cur_row, column=real_col)
            cell.number_format = '#,##0.00'
            if md['status'] == 'ok':
                cell.fill = ok_fill; cell.font = wht_font
            elif md['status'] == 'over':
                cell.fill = over_fill; cell.font = wht_font
            elif md['status'] == 'under':
                cell.fill = under_fill; cell.font = blk_font
            meta_cell = ws.cell(row=cur_row, column=real_col - 1)
            meta_cell.number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 10
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=Orcamento_{active_company.name}_{year}.xlsx'
    )
    wb.save(response)
    return response


# Mantém compatibilidade com budget_edit_view antigo (redireciona para nova página)
@login_required
def budget_edit_view(request):
    return redirect('core:budget_entry')
